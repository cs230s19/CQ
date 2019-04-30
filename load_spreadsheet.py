import json

from openpyxl import load_workbook


def get_spreadsheet(filename):
    """
    This function simply loads the excel sheet and returns it as a workbook
    If the file cannot be read then an error message is printed
    :param filename: name of the excel file
    :return: excel file as a workbook
    """
    try:
        workbook = load_workbook(filename)
    except Exception as e:
        print("Unable to load file: {}".format(e))
        return None

    return workbook


def get_actions(filename):
    """
    Parses through the excel sheet and creates a json string containing the attributes
    of each action in the spreadsheet

    :param filename: name of spreadsheet
    :return: json string
    """
    workbook = get_spreadsheet(filename)

    if workbook is not None:
        # Use first sheet
        worksheet = workbook[workbook.sheetnames[0]]

        headers = []
        actions = []

        first_row = tuple(worksheet.rows)[0]
        for cell in first_row:
            if cell.value is not None:
                headers.append(cell.value.strip())

        for row in worksheet.iter_rows(min_row=2, max_col=len(headers)):
            action = {'Times Completed': 0}
            for i in range(len(headers)):
                action[str(headers[i])] = str(row[i].value)

            actions.append(action)

        return json.dumps(actions)


if __name__ == '__main__':
    parsed = json.loads(get_actions("CQ Point System.xlsx"))

    print(json.dumps(parsed, indent=4, sort_keys=True))
